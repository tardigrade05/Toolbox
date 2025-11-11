from openpyxl import load_workbook
from copy import copy
import datetime
import shutil
import joRmaHelper as helper

def insertToExcelRMA(constant_list,item_list):
    try:
        number_cell = []
        item_name_cell = []
        description_cell = []
        qty_cell = []
        serial_num_cell = []
        remarks_cell = []
        dop_cell = []
        dr_num_cell = []
        uniqueCode= ""
    
        year =str(datetime.datetime.now().strftime("%Y"))
        
        path = "Source/RMA/RMAForm.xlsx"
        ref_workBook = load_workbook(path)
        sheet=ref_workBook.active
        
        if len(item_list) != 1:
            sheet.insert_rows(9,len(item_list)-1)
            for i in range(len(item_list)-1):
                copy_row_format(sheet, 9 + len(item_list)-1, 9 + i)
            
        branch_cell = sheet['C4']
        client_name_cell = sheet['C5']
        contact_num_cell = sheet['C6']
        reference_num_cell = sheet['G5']
        date_cell = sheet['G6']
        
        #item_set
        for index in range(0,len(item_list)):
            # if len(item_list) != 1:
            number_cell.append(sheet[f'B{9+index}'])
            item_name_cell.append(sheet[f'C{9+index}'])
            description_cell.append(sheet[f'D{9+index}'])
            qty_cell.append(sheet[f'E{9+index}'])
            serial_num_cell.append(sheet[f'F{9+index}'])
            remarks_cell.append(sheet[f'G{9+index}'])
            dop_cell.append(sheet[f'H{9+index}'])
            dr_num_cell.append(sheet[f'I{9+index}'])
            branchRecievedBy = sheet[f'D{11+index}']
            tech_name_cell = sheet[f'H{11+index}']
            branchRecievedByDate = sheet[f'D{12+index}']
            techRecievedByDate = sheet[f'H{12+index}']

        branch_cell.value = constant_list[0]
        client_name_cell.value = str(constant_list[2]).upper()
        contact_num_cell.value = constant_list[3]
        reference_num_cell.value = constant_list[4]
        date_cell.value = helper.formatDate("/","dd/mm/yyyy",constant_list[1])
        branchRecievedBy.value = str(constant_list[9]).upper()
        branchRecievedByDate.value = helper.formatDate("/","dd/mm/yyyy",constant_list[1])
        tech_name_cell.value = str(constant_list[9]).upper()
        techRecievedByDate.value = helper.formatDate("/","dd/mm/yyyy",constant_list[1])

        for index,value in zip(range(0,len(item_list)),item_list):
            number_cell[index].value = index+1
            item_name_cell[index].value = value[0].get()
            description_cell[index].value = value[1].get()
            qty_cell[index].value = 1
            serial_num_cell[index].value = "*"+str(value[2].get())
            remarks_cell[index].value = str(value[4].get()).upper()
            dop_cell[index].value = helper.formatDate("/","dd/mm/yyyy",constant_list[6])
            dr_num_cell[index].value = "*"+str(constant_list[10])

        uniqueCode = f"{helper.formatDate("-","mm/dd/yyyy",constant_list[1])}_{constant_list[10]}{constant_list[4]}"
        ref_workBook.save(filename=f"Data/RMA/{str(year)}/{str(constant_list[2]).upper()}-RMA {uniqueCode}.xlsx")
        shutil.copy(f"Data/RMA/{str(year)}/{str(constant_list[2]).upper()}-RMA {uniqueCode}.xlsx",f"{helper.targetPath("Desktop")}/JO_RMA/Data/RMA/{str(year)}/")
        filename = f"{helper.targetPath("Desktop")}/JO_RMA/Data/RMA/{str(year)}/{str(constant_list[2]).upper()}-RMA {uniqueCode}.xlsx"
        helper.write_txt(filename,"filename.txt",'w')

    except Exception as e:
        print(f'Rma Error {e}')
        return
    
def insertToExcelJO(constant_list,item_list):
    try:
        qty_cell = []
        item_code_cell = []
        description_cell = []
        serial_num_cell = []
        customer_complaint_cell = []
        tech_finding_cell = []
        uniqueCode= ""

        year =str(datetime.datetime.now().strftime("%Y"))
        date = str(datetime.datetime.now().strftime("%m-%d-%Y_"))
        
        path = "Source/JO/JobOrderForm.xlsx"
        ref_workBook = load_workbook(path)
        sheet=ref_workBook.active

        if len(item_list) != 1:
            sheet.insert_rows(24,len(item_list)-1)
            for i in range(len(item_list)-1):
                copy_row_format(sheet, 24 + len(item_list)-1, 24 + i)

        dr_num_cell = sheet['D8']
        customer_name_cell = sheet['D9']
        supplier_name_cell = sheet['D10']
        jobOrder_num_cell = sheet['H6']
        check_in_cell = sheet['H7']
        warranty_start_cell = sheet['H8']
        warranty_end_cell = sheet['H9']
        remaining_warranty_cell = sheet['H10']
        customer_name2_cell = sheet['F19']
        
        
        for index in range(len(item_list)):
            qty_cell.append(sheet[f'C{24+index}'])
            item_code_cell.append(sheet[f'D{24+index}'])
            description_cell.append(sheet[f'E{24+index}'])
            serial_num_cell.append(sheet[f'F{24+index}'])
            customer_complaint_cell.append(sheet[f'G{24+index}'])
            tech_finding_cell.append(sheet[f'H{24+index}'])
            recieving_staff_cell = sheet[f'E{26+index}']
        
        dr_num_cell.value = "*"+str(constant_list[10])
        customer_name_cell.value = str(constant_list[2]).upper()
        supplier_name_cell.value = "JOYO MARKETING"
        jobOrder_num_cell.value = constant_list[5]
        check_in_cell.value = helper.formatDate("/","dd/mm/yyyy",constant_list[1])
        warranty_start_cell.value = helper.formatDate("/","dd/mm/yyyy",constant_list[6])
        warranty_end_cell.value = helper.formatDate("/","dd/mm/yyyy",constant_list[7])
        remaining_warranty_cell.value = constant_list[8]
        customer_name2_cell.value = str(constant_list[2]).upper()
        recieving_staff_cell.value = str(constant_list[9]).upper()

        for index,value in zip(range(0,len(item_list)),item_list):
            qty_cell[index].value = 1
            item_code_cell[index].value = value[0].get()
            description_cell[index].value = value[1].get()
            serial_num_cell[index].value = value[2].get()
            customer_complaint_cell[index].value = str(value[3].get()).upper()
            tech_finding_cell[index].value = str(value[4].get()).upper()
            changeRowColumnCell(sheet,"E",f"{24+index}",value[1].get())
        
        merge_cell(sheet,len(item_list)-1,3,8)

        uniqueCode = f"{helper.formatDate("-","mm/dd/yyyy",constant_list[1])}_{constant_list[10]}{constant_list[5][-1]}"
        ref_workBook.save(filename=f"Data/JO/{str(year)}/{str(constant_list[2]).upper()}-JO {uniqueCode}.xlsx")
        shutil.copy(f"Data/JO/{str(year)}/{str(constant_list[2]).upper()}-JO {uniqueCode}.xlsx",f"{helper.targetPath("Desktop")}/JO_RMA/Data/JO/{str(year)}/")

    except Exception as e:
        print(f'JO Error {e}')
        return


def changeRowColumnCell(sheet,column,row,value):
    textLength = len(str(value))
    heightTimes = int(textLength/20)
    if heightTimes < 1:
        sheet.row_dimensions[int(row)].height = 20
        sheet.column_dimensions[column].width = 30
    else:
        sheet.row_dimensions[int(row)].height = 20*(heightTimes+1)
        sheet.column_dimensions[column].width = 30
    # save the file 

# def insertRMANumberColor(column,row):
#     path = "Source/RMA/RMANumber.xlsx"
#     ref_workBook = openpyxl.load_workbook(path)
#     sheet = ref_workBook.active
#     sheet[f'{column}{row}'].fill = PatternFill(patternType='solid',start_color='FFFF00')
#     ref_workBook.save(filename=path)


# def insertJONumber(column,row,value):
#     path = "Source/JO/JobOrderNumber.xlsx"
#     ref_workBook = openpyxl.load_workbook(path)
#     sheet = ref_workBook.active
    

#     inputCell = sheet[f'{column}{row}']
#     inputCell.value = value
    

#     ref_workBook.save(filename=path)

# def rmaCheckCell(value):
#     path = "Source/RMA/RMANumber.xlsx"
#     ref_workBook = openpyxl.load_workbook(path)
#     sheet = ref_workBook.active
#     coordinate = (DB.showData()[-1][0]).split("-")
#     row = coordinate[1] #5
#     column = coordinate[0] #G
   
# #sheet[f"{chr(ord(column)-1)}{int(row)+1}"].value
#     #print(str(sheet[f"{chr(ord(column)+1)}{int(row)}"].value))
#     #print(str(value))


#     if str(sheet[f"{chr(ord(column)+1)}{int(row)}"].value) ==str(value):
#         row = str(int(row))
#         column = chr(ord(column)+1)
#         insertRMANumberColor(column,row)
#         #print("Rma nextline...")
#         return str(column)+"-"+str(row)
#     elif str(sheet[f"{chr(ord(column)+1)}{int(row)}"].value) !=str(value):
#         row = str(int(row)+1)
#         column = "A"
#         insertRMANumberColor(column,row)
#         #print("Rma newline...")
#         return str(column)+"-"+str(row)
#     else:
#         #print("ERROR OCCURS...")
#         return'A0'

# def joCheckCell(value):
#     path = "Source/JO/JobOrderNumber.xlsx"
#     ref_workBook = openpyxl.load_workbook(path)
#     sheet = ref_workBook.active
#     coordinate = (DB.showData()[-1][1]).split("-")
#     row = coordinate[1] #23s
#     column = coordinate[0] #B
#     #print("This is the Jo Cell Value "+str(sheet[f"{chr(ord(column)-1)}{int(row)+1}"].value))

#     if int(ord(value[-1])) >90:
#         value[-1] = "A"

#     if str(sheet[f"{chr(ord(column)-1)}{int(row)+1}"].value) !="None":
#         row = str(int(row)+1)
#         column = chr(ord(column))
#         insertJONumber(column,row,value)
#         #print("Jo nextline...")
#         return str(column)+"-"+str(row)
#     elif str(sheet[f"{chr(ord(column)-1)}{int(row)+1}"].value) =="None" or str(sheet[f"{chr(ord(column)-1)}{int(row)+1}"].value) =="" or str(sheet[f"{chr(ord(column)-1)}{int(row)+1}"].value) ==" ":
#         row = "2"
#         column = chr(ord(column)+2)
#         insertJONumber(column,row,value)
#         #print("Jo newline...")
#         return str(column)+"-"+str(row)
#     else:
#         #print("ERROR OCCURS...")
#         return'A0'
    
#     #return'A', "0"



def copy_row_format(ws, source_row, new_row):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        new_cell = ws.cell(row=new_row, column=col)
        new_cell.font = copy(source_cell.font)
        new_cell.border = copy(source_cell.border)
        new_cell.fill= copy(source_cell.fill)
        new_cell.alignment = copy(source_cell.alignment)
        new_cell.number_format = copy(source_cell.number_format)
        new_cell.protection = copy(source_cell.protection)


def merge_cell(ws,item_length,c_from,c_end):
    row_list = [29,36,43,49]
    for value in row_list:
        ws.merge_cells(start_row=item_length+value, start_column=c_from, end_row=item_length+value, end_column=c_end)
